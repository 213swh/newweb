
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

def match_and_insert_data(source_file, target_file, output_file):
    """
    将source_file的第一列和第二列作为主键，与target_file的第二列和第三列作为主键进行匹配
    不匹配的数据插入到target_file中，并设置特定字段值
    
    参数:
        source_file: 源Excel文件路径
        target_file: 目标Excel文件路径
        output_file: 输出Excel文件路径
    """
    # 加载工作簿
    source_wb = load_workbook(filename=source_file)
    target_wb = load_workbook(filename=target_file)
    
    # 假设两个文件都只有一个工作表，或者我们处理第一个工作表
    source_ws = source_wb.active
    target_ws = target_wb.active
    
    # 获取target文件中的主键集合（第二列和第三列的组合）
    target_keys = set()
    for row in range(2, target_ws.max_row + 1):  # 假设第一行是标题行
        col2 = target_ws.cell(row=row, column=2).value
        col3 = target_ws.cell(row=row, column=3).value
        if col2 is not None and col3 is not None:
            key = (str(col2).strip(), str(col3).strip())
            target_keys.add(key)
    
    # 遍历source文件，查找不匹配的行
    new_rows = []
    for row in range(2, source_ws.max_row + 1):  # 假设第一行是标题行
        col1 = source_ws.cell(row=row, column=1).value
        col2 = source_ws.cell(row=row, column=2).value
        
        if col1 is None or col2 is None:
            continue
            
        source_key = (str(col1).strip(), str(col2).strip())
        
        # 如果主键不在target文件中，则收集这一行数据
        if source_key not in target_keys:
            # 获取整行数据
            row_data = []
            for col in range(1, source_ws.max_column + 1):
                row_data.append(source_ws.cell(row=row, column=col).value)
            
            new_rows.append(row_data)
    
    # 将不匹配的行插入到target文件中
    for row_data in new_rows:
        # 确定插入位置（最后一行之后）
        insert_row = target_ws.max_row + 1
        
        # 插入源数据（注意列对应关系）
        # 假设需要将source的列1,2对应到target的列2,3
        # 这里需要根据实际情况调整列映射关系
        target_ws.cell(row=insert_row, column=2, value=row_data[0])  # source列1 -> target列2
        target_ws.cell(row=insert_row, column=3, value=row_data[1])  # source列2 -> target列3
        
        # 插入其他列数据（如果有需要）
        # 这里可以根据实际情况添加更多列的数据复制
        
        # 设置特定字段值
        target_ws.cell(row=insert_row, column=9, value="N")  # 第九个字段"有上报"设为N
        target_ws.cell(row=insert_row, column=10, value="Y")  # 第十个字段"有设计"设为Y
    
    # 保存处理后的工作簿
    target_wb.save(output_file)
    print(f"处理完成！结果已保存至: {output_file}")

# 使用示例
if __name__ == "__main__":
    source_file = "/Users/handing1993/Desktop/22.xlsx"  # 源文件路径
    target_file = "/Users/handing1993/Desktop/merged_result66.xlsx"  # 目标文件路径
    output_file = "/Users/handing1993/Desktop/final_result.xlsx"  # 输出文件路径
    
    match_and_insert_data(source_file, target_file, output_file)